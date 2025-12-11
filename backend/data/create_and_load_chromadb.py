module_description = r"""ChromaDB Document Management for Selection Descriptions

This module provides functionality to manage selection descriptions in ChromaDB,
with support for document deduplication, embedding generation, and similarity search.

Key Features:
-------------
1. Document Management:
   - SQLite to ChromaDB document transfer
   - Document deduplication using MD5 hashing
   - Metadata preservation (selection codes)
   - UUID-based document identification
   - Automatic document splitting for long texts
   - Token counting and validation

2. Embedding Generation:
   - Azure OpenAI embedding model integration
   - Batch embedding generation
   - Configurable model deployment
   - Token limit handling (8190 tokens max)
   - Smart text chunking for long documents

3. Similarity Search:
   - Query embedding generation
   - Configurable result count
   - Distance-based ranking
   - Metadata and document retrieval
   - Support for split document reconstruction

4. Error Handling:
   - Database connection management
   - Embedding generation error handling
   - Document validation
   - Debug logging support
   - Token limit error handling

5. Performance:
   - Batch processing
   - Efficient deduplication
   - Persistent storage
   - Connection pooling
   - Smart document chunking

Processing Flow:
--------------
1. Initialization:
   - Sets up project paths
   - Configures database connections
   - Initializes embedding client
   - Sets up debug logging

2. Document Retrieval:
   - Connects to SQLite database
   - Retrieves documents and selection codes
   - Generates document hashes
   - Validates document content
   - Counts tokens for each document

3. Document Processing:
   - Checks for existing documents
   - Filters out duplicates
   - Generates unique IDs
   - Prepares metadata
   - Splits long documents if needed
   - Validates token counts

4. Embedding Generation:
   - Batches documents for processing
   - Generates embeddings using Azure
   - Handles API responses
   - Validates embedding results
   - Processes document chunks

5. ChromaDB Integration:
   - Creates/connects to collection
   - Adds documents with embeddings
   - Stores metadata
   - Updates collection
   - Maintains chunk relationships

6. Search Capabilities:
   - Query embedding generation
   - Similarity search execution
   - Result ranking and formatting
   - Metadata retrieval
   - Chunk reconstruction

Usage Example:
-------------
# Initialize and populate ChromaDB
collection = upsert_documents_to_chromadb(
    deployment="text-embedding-3-small_mimi",
    collection_name="my_collection"
)

# Perform similarity search
results = collection.query(
    query_embeddings=[query_embedding],
    n_results=3,
    include=["documents", "metadatas", "distances"]
)

Required Environment:
-------------------
- Python 3.7+
- Azure OpenAI API access
- SQLite database with selection descriptions
- Write permissions for ChromaDB directory
- tiktoken package for token counting

Output:
-------
- ChromaDB collection with:
  - Document embeddings
  - Selection code metadata
  - Document hashes
  - Unique document IDs
  - Chunk information for split documents

Error Handling:
-------------
- Database connection errors
- Embedding generation failures
- Document validation errors
- API rate limiting
- File system errors
- Token limit errors
- Chunk processing errors"""

import hashlib
import logging

# ==============================================================================
# IMPORTS
# ==============================================================================
# Standard library imports
import os
import re
import sqlite3
import sys
import time
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple
from uuid import uuid4
import math

# Third-party imports
import chromadb
import numpy as np
import openpyxl
import tiktoken
import tqdm as tqdm_module
from langchain_community.retrievers import BM25Retriever
from langchain_core.documents import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Load environment variables from .env file
from dotenv import load_dotenv

load_dotenv()

# Add this import after the other imports
try:
    from rank_bm25 import BM25Okapi

    print("rank_bm25 is available. BM25 search will be enabled.")
except ImportError:
    print("Warning: rank_bm25 not available. BM25 search will be disabled.")
    BM25Okapi = None

# ==============================================================================
# PATH SETUP
# ==============================================================================
# --- Ensure project root is in sys.path for local imports ---
try:
    BASE_DIR = Path(__file__).resolve().parents[1]
except NameError:
    BASE_DIR = Path(os.getcwd()).parents[0]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

# Direct implementation of model functions to avoid circular imports completely
from openai import AzureOpenAI
from langchain_openai import AzureOpenAIEmbeddings
from api.utils.debug import print__chromadb_debug


# ===============================================================================
# Azure Embedding Models
# ===============================================================================
def get_azure_embedding_model():
    """Get an instance of Azure OpenAI Embedding model with standard configuration.

    Returns:
        AzureOpenAI: Configured embedding client instance
    """
    return AzureOpenAI(
        api_version="2024-12-01-preview",
        azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
        api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    )


def get_langchain_azure_embedding_model(model_name="text-embedding-3-small_mimi"):
    """Get a LangChain AzureOpenAIEmbeddings instance with standard configuration.

    Args:
        model_name (str): The name of the embedding model deployment

    Returns:
        AzureOpenAIEmbeddings: Configured embedding model instance
    """
    from langchain_openai import AzureOpenAIEmbeddings

    return AzureOpenAIEmbeddings(
        model=model_name,
        azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
        api_key=os.getenv("AZURE_OPENAI_API_KEY"),
        openai_api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview"),
        deployment=model_name,
    )


# ==============================================================================
# CONSTANTS & CONFIGURATION
# ==============================================================================
# Database paths
CHROMA_DB_PATH = BASE_DIR / "metadata" / "czsu_chromadb"
SQLITE_DB_PATH = (
    BASE_DIR / "metadata" / "llm_selection_descriptions" / "selection_descriptions.db"
)

# Unique identifier for this module's debug messages
CREATE_CHROMADB_ID = 30

# Token limit for Azure OpenAI
MAX_TOKENS = 8190

# SQL query for retrieving documents
SELECT_DOCUMENTS_QUERY = (
    "SELECT extended_description, selection_code " "FROM selection_descriptions"
)


# ==============================================================================
# MONITORING AND METRICS
# ==============================================================================
@dataclass
class Metrics:
    """Simple metrics collection for tracking processing statistics.

    This class tracks various metrics during the processing of documents:
    - Processing time
    - Success/failure counts
    - Failed documents with reasons

    Attributes:
        start_time (float): Timestamp when processing started.
        processed_docs (int): Number of successfully processed documents.
        failed_docs (int): Number of documents that failed processing.
        total_processing_time (float): Total time taken for processing.
        failed_records (list): List of tuples containing (selection_code, error_message) for failed records.
    """

    start_time: float = field(default_factory=time.time)
    processed_docs: int = 0
    failed_docs: int = 0
    total_processing_time: float = 0
    failed_records: list = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        """Convert metrics to a dictionary.

        Returns:
            Dict[str, Any]: A dictionary containing the metrics with formatted timestamps
                           and calculated averages.
        """
        return {
            "start_time": datetime.fromtimestamp(self.start_time).isoformat(),
            "processed_docs": self.processed_docs,
            "failed_docs": self.failed_docs,
            "total_processing_time": self.total_processing_time,
            "average_time_per_doc": self.total_processing_time
            / max(1, self.processed_docs),
            "success_rate": (
                self.processed_docs / max(1, self.processed_docs + self.failed_docs)
            )
            * 100,
            "failed_records": self.failed_records,
        }

    def update_processing_time(self) -> None:
        """Update the total processing time based on the current time."""
        self.total_processing_time = time.time() - self.start_time


def handle_processing_error(
    error: Exception, selection_code: str, metrics: Metrics
) -> None:
    """Handle processing errors consistently.

    This function provides consistent error handling by:
    1. Formatting error messages uniformly
    2. Logging to both console and file
    3. Updating metrics
    4. Ensuring proper error propagation

    Args:
        error (Exception): The error that occurred.
        selection_code (str): The selection code being processed.
        metrics (Metrics): The metrics object to update.
    """
    error_msg = f"Error processing selection code {selection_code}: {str(error)}"
    print__chromadb_debug(f"\n{error_msg}")
    metrics.failed_docs += 1
    metrics.failed_records.append((selection_code, str(error)))


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================
def get_document_hash(text: str) -> str:
    """Generate MD5 hash for a document text.

    This function creates a unique hash for each document to enable
    efficient deduplication. The hash is generated using MD5 and
    is based on the UTF-8 encoded text content.

    Args:
        text (str): The document text to hash

    Returns:
        str: MD5 hash of the document text

    Raises:
        UnicodeEncodeError: If the text cannot be encoded as UTF-8
    """
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def get_documents_from_sqlite() -> tuple[list[str], list[str], list[str]]:
    """Retrieve documents from SQLite database.

    This function connects to the SQLite database and retrieves all
    documents and their corresponding selection codes. It also
    generates MD5 hashes for each document to enable deduplication.

    Returns:
        tuple: (texts, selections, hashes) where:
            - texts is a list of document contents
            - selections is a list of corresponding selection codes
            - hashes is a list of MD5 hashes of the documents

    Raises:
        sqlite3.Error: If there's an error connecting to or querying the database
    """
    try:
        conn = sqlite3.connect(str(SQLITE_DB_PATH))
        cursor = conn.cursor()

        # First check if the table exists
        cursor.execute(
            """
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='selection_descriptions'
        """
        )
        if not cursor.fetchone():
            print__chromadb_debug(
                f"⚠️ {CREATE_CHROMADB_ID}: Table 'selection_descriptions' does not exist in SQLite database."
            )
            return [], [], []

        # Get only documents that have extended_description
        cursor.execute(
            """
            SELECT extended_description, selection_code 
            FROM selection_descriptions 
            WHERE extended_description IS NOT NULL 
            AND extended_description != ''
        """
        )
        results = cursor.fetchall()

        if not results:
            print__chromadb_debug(
                f"⚠️ {CREATE_CHROMADB_ID}: No documents found in SQLite database."
            )
            return [], [], []

        texts, selections = zip(*results)
        hashes = [get_document_hash(text) for text in texts]

        # Print some debug info about the documents
        print__chromadb_debug(
            f"📊 {CREATE_CHROMADB_ID}: Found {len(texts)} documents in SQLite database."
        )
        print__chromadb_debug(f"📊 {CREATE_CHROMADB_ID}: Sample document lengths:")
        for sel, text in list(zip(selections, texts))[:5]:
            print__chromadb_debug(
                f"📊 {CREATE_CHROMADB_ID}: - {sel}: {len(text)} characters"
            )

        return list(texts), list(selections), list(hashes)

    except sqlite3.Error as e:
        print__chromadb_debug(f"❌ {CREATE_CHROMADB_ID}: Database error: {str(e)}")
        raise
    finally:
        if "conn" in locals():
            conn.close()


def num_tokens_from_string(string: str, encoding_name: str = "cl100k_base") -> int:
    """Count the number of tokens in a string using tiktoken.

    Args:
        string (str): The text to count tokens for
        encoding_name (str): The encoding to use (default: cl100k_base for text-embedding-3-large)

    Returns:
        int: Number of tokens in the string
    """
    encoding = tiktoken.get_encoding(encoding_name)
    return len(encoding.encode(string))


def split_text_by_tokens(text: str, max_tokens: int = MAX_TOKENS) -> List[str]:
    """Split text into chunks that don't exceed the token limit, using token-based splitting.

    Args:
        text (str): The text to split
        max_tokens (int): Maximum tokens per chunk

    Returns:
        List[str]: List of text chunks, each under the token limit
    """
    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(text)
    total_tokens = len(tokens)
    if total_tokens <= max_tokens:
        return [text]

    num_chunks = math.ceil(float(total_tokens) / max_tokens)
    chunks = []
    for i in range(num_chunks):
        start = i * max_tokens
        end = min((i + 1) * max_tokens, total_tokens)
        chunk_tokens = tokens[start:end]
        chunk_text = encoding.decode(chunk_tokens)
        chunks.append(chunk_text)
    return chunks


def normalize_czech_text(text: str) -> str:
    """Advanced Czech text normalization for better search matching."""
    if not text:
        return text

    # Convert to lowercase first
    text = text.lower()

    # Advanced Czech diacritics mapping for normalization
    czech_diacritics_map = {
        # Primary Czech diacritics
        "á": "a",
        "č": "c",
        "ď": "d",
        "é": "e",
        "ě": "e",
        "í": "i",
        "ň": "n",
        "ó": "o",
        "ř": "r",
        "š": "s",
        "ť": "t",
        "ú": "u",
        "ů": "u",
        "ý": "y",
        "ž": "z",
        # Extended mappings for robustness
        "à": "a",
        "ä": "a",
        "â": "a",
        "ă": "a",
        "ą": "a",
        "ç": "c",
        "ć": "c",
        "ĉ": "c",
        "ċ": "c",
        "è": "e",
        "ë": "e",
        "ê": "e",
        "ę": "e",
        "ė": "e",
        "ē": "e",
        "ì": "i",
        "ï": "i",
        "î": "i",
        "į": "i",
        "ī": "i",
        "ò": "o",
        "ö": "o",
        "ô": "o",
        "õ": "o",
        "ő": "o",
        "ø": "o",
        "ù": "u",
        "ü": "u",
        "û": "u",
        "ű": "u",
        "ū": "u",
        "ÿ": "y",
        "ŷ": "y",
        "ź": "z",
        "ż": "z",
    }

    # Create ASCII version
    ascii_text = text
    for diacritic, ascii_char in czech_diacritics_map.items():
        ascii_text = ascii_text.replace(diacritic, ascii_char)

    # Return both versions separated by space for broader indexing
    if ascii_text != text:
        return f"{text} {ascii_text}"
    return text


def hybrid_search(
    collection, query_text: str, n_results: int = 60, rare_terms: Set[str] = None
) -> List[Dict]:
    """
    Hybrid search that combines semantic and BM25 approaches with semantic focus.

    This function trusts text-embedding-3-large's semantic capabilities while using
    BM25 for exact keyword matches. The approach is semantic-focused, meaning:
    - Semantic search gets higher weight (0.85) as the primary method
    - BM25 search gets lower weight (0.15) for exact matches only
    - Results are combined and ranked by weighted score

    Args:
        collection: ChromaDB collection to search
        query_text: The search query string
        n_results: Maximum number of results to return (default: 60)
        rare_terms: Set of rare terms (unused, kept for compatibility)

    Returns:
        List[Dict]: Ranked search results with metadata including:
            - document: The document content
            - metadata: Document metadata (selection codes, etc.)
            - score: Final weighted score
            - semantic_score: Normalized semantic similarity score
            - bm25_score: Normalized BM25 relevance score
            - source: Source of the result ('semantic', 'bm25', 'hybrid', 'fallback_semantic')
    """

    logging.info(f"Hybrid search for query: '{query_text}'")

    try:
        # Step 1: Clean and normalize query (minimal processing)
        normalized_query = normalize_czech_text(query_text)

        # Step 2: Perform semantic search (primary method)
        semantic_results = []
        try:
            embedding_client = get_azure_embedding_model()

            semantic_raw = similarity_search_chromadb(
                collection=collection,
                embedding_client=embedding_client,
                query=normalized_query,
                embedding_model_name="text-embedding-3-small_mimi",
                k=n_results,
            )

            for i, (doc, meta, distance) in enumerate(
                zip(
                    semantic_raw["documents"][0],
                    semantic_raw["metadatas"][0],
                    semantic_raw["distances"][0],
                )
            ):
                # Convert distance to similarity score
                similarity_score = max(0, 1 - (distance / 2))

                semantic_results.append(
                    {
                        "id": f"semantic_{i}",
                        "document": doc,
                        "metadata": meta,
                        "semantic_score": similarity_score,
                        "source": "semantic",
                    }
                )

            logging.info(f"Semantic search returned {len(semantic_results)} results")

        except Exception as e:
            logging.error(f"Semantic search failed: {e}")
            semantic_results = []

        # Step 3: Perform minimal BM25 search (for exact keyword matches)
        bm25_results = []
        try:
            all_data = collection.get(include=["documents", "metadatas"])

            if all_data and "documents" in all_data and all_data["documents"]:
                documents = all_data["documents"]
                metadatas = all_data["metadatas"]

                # Simple document processing - just normalize
                processed_docs = [normalize_czech_text(doc) for doc in documents]

                if BM25Okapi:
                    tokenized_docs = [doc.split() for doc in processed_docs]
                    bm25 = BM25Okapi(tokenized_docs)

                    # Simple query processing
                    tokenized_query = normalized_query.split()
                    bm25_scores = bm25.get_scores(tokenized_query)

                    # Get top results
                    top_indices = np.argsort(bm25_scores)[::-1][:n_results]

                    for i, idx in enumerate(top_indices):
                        if bm25_scores[idx] > 0:
                            bm25_results.append(
                                {
                                    "id": f"bm25_{i}",
                                    "document": documents[idx],
                                    "metadata": (
                                        metadatas[idx] if idx < len(metadatas) else {}
                                    ),
                                    "bm25_score": float(bm25_scores[idx]),
                                    "source": "bm25",
                                }
                            )

                    logging.info(f"BM25 search returned {len(bm25_results)} results")

        except Exception as e:
            logging.error(f"BM25 search failed: {e}")
            bm25_results = []

        # Step 4: Combine results with semantic-focused weighting
        combined_results = {}

        # Process semantic results (primary)
        for result in semantic_results:
            doc_id = result["metadata"].get("selection", result["document"][:50])
            if doc_id not in combined_results:
                combined_results[doc_id] = result.copy()
                combined_results[doc_id]["bm25_score"] = 0.0

        # Process BM25 results (secondary)
        for result in bm25_results:
            doc_id = result["metadata"].get("selection", result["document"][:50])
            if doc_id in combined_results:
                combined_results[doc_id]["bm25_score"] = result["bm25_score"]
                combined_results[doc_id]["source"] = "hybrid"
            else:
                combined_results[doc_id] = result.copy()
                combined_results[doc_id]["semantic_score"] = 0.0

        # Step 5: Calculate final scores with semantic focus
        final_results = []
        max_semantic = max(
            (r.get("semantic_score", 0) for r in combined_results.values()), default=1
        )
        max_bm25 = max(
            (r.get("bm25_score", 0) for r in combined_results.values()), default=1
        )

        # Semantic-focused weights: trust the embedding model more
        semantic_weight = 0.85  # High weight for semantic
        bm25_weight = 0.15  # Low weight for exact matches only

        for doc_id, result in combined_results.items():
            # Normalize scores
            semantic_score = (
                result.get("semantic_score", 0.0) / max_semantic
                if max_semantic > 0
                else 0.0
            )
            bm25_score = (
                result.get("bm25_score", 0.0) / max_bm25 if max_bm25 > 0 else 0.0
            )

            # Calculate final score with semantic focus
            final_score = (semantic_weight * semantic_score) + (
                bm25_weight * bm25_score
            )

            result["score"] = final_score
            result["semantic_score"] = semantic_score
            result["bm25_score"] = bm25_score
            result["weights_used"] = {"semantic": semantic_weight, "bm25": bm25_weight}

            final_results.append(result)

        # Sort by final score
        final_results.sort(key=lambda x: x["score"], reverse=True)

        # Return top results
        top_results = final_results[:n_results]
        logging.info(f"Hybrid search completed, returning {len(top_results)} results")

        # Log top result details for debugging
        if top_results:
            top = top_results[0]
            logging.info(
                f"Top result: {top['metadata'].get('selection', 'unknown')} "
                f"(score: {top['score']:.4f}, semantic: {top['semantic_score']:.4f}, "
                f"bm25: {top['bm25_score']:.4f})"
            )

        return top_results

    except Exception as e:
        logging.error(f"Hybrid search failed: {e}")

        # Fallback to pure semantic search
        try:
            embedding_client = get_azure_embedding_model()
            fallback_results = similarity_search_chromadb(
                collection=collection,
                embedding_client=embedding_client,
                query=query_text,
                embedding_model_name="text-embedding-3-small_mimi",
                k=n_results,
            )

            converted_results = []
            for i, (doc, meta, distance) in enumerate(
                zip(
                    fallback_results["documents"][0],
                    fallback_results["metadatas"][0],
                    fallback_results["distances"][0],
                )
            ):
                similarity_score = max(0, 1 - (distance / 2))
                converted_results.append(
                    {
                        "id": f"fallback_{i}",
                        "document": doc,
                        "metadata": meta,
                        "score": similarity_score,
                        "semantic_score": similarity_score,
                        "bm25_score": 0.0,
                        "source": "fallback_semantic",
                    }
                )

            return converted_results

        except Exception as fallback_error:
            logging.error(f"Fallback search also failed: {fallback_error}")
            return []


# ==============================================================================
# MAIN LOGIC
# ==============================================================================
def upsert_documents_to_chromadb(
    deployment: str = "text-embedding-3-small_mimi",
    collection_name: str = "czsu_selections_chromadb",
) -> chromadb.Collection | None:
    """Add documents from SQLite to a ChromaDB collection.

    This function:
    1. Retrieves documents from SQLite
    2. Checks for existing documents in ChromaDB
    3. Generates embeddings for new documents
    4. Adds new documents to ChromaDB with metadata

    Args:
        deployment (str): Azure embedding deployment name to use
        collection_name (str): Name of the ChromaDB collection to use or create

    Returns:
        chromadb.Collection | None: The ChromaDB collection object after updates,
                                   or None if no documents were found

    Raises:
        ValueError: If no documents are found in SQLite
        chromadb.errors.ChromaDBError: If there's an error with ChromaDB operations
        Exception: For other unexpected errors
    """
    metrics = Metrics()

    try:
        # Get documents from SQLite
        print(f"🔍 {CREATE_CHROMADB_ID}: Getting documents from SQLite...")
        texts, selections, hashes = get_documents_from_sqlite()
        print(f"📊 {CREATE_CHROMADB_ID}: Retrieved {len(texts)} documents from SQLite")
        if not texts:
            print__chromadb_debug(
                f"⚠️ {CREATE_CHROMADB_ID}: No documents found in SQLite database."
            )
            return None

        # Initialize Azure embedding client
        print(f"🔑 {CREATE_CHROMADB_ID}: Initializing Azure embedding client...")
        embedding_client = get_azure_embedding_model()

        # Initialize ChromaDB client and get/create collection with cloud/local support
        print(f"📦 {CREATE_CHROMADB_ID}: Initializing ChromaDB client...")
        from metadata.chromadb_client_factory import get_chromadb_client

        try:
            client = get_chromadb_client(
                local_path=CHROMA_DB_PATH, collection_name=collection_name
            )
            print(f"✅ {CREATE_CHROMADB_ID}: ChromaDB client initialized successfully")
        except Exception as client_error:
            print(
                f"❌ {CREATE_CHROMADB_ID}: Error initializing ChromaDB client: {client_error}"
            )
            import traceback

            traceback.print_exc()
            raise

        try:
            # Try to create the collection with cosine similarity if it doesn't exist
            print(
                f"🆕 {CREATE_CHROMADB_ID}: Creating new collection '{collection_name}'..."
            )
            collection = client.create_collection(
                name=collection_name, metadata={"hnsw:space": "cosine"}
            )
            print(f"✅ {CREATE_CHROMADB_ID}: Collection created successfully")
        except Exception as coll_error:
            # If it already exists, just get it
            print(
                f"ℹ️ {CREATE_CHROMADB_ID}: Collection exists, retrieving it... ({coll_error})"
            )
            collection = client.get_collection(name=collection_name)
            print(f"✅ {CREATE_CHROMADB_ID}: Collection retrieved successfully")

        # Check for existing documents in ChromaDB
        existing = collection.get(include=["metadatas"], limit=10000)
        existing_hashes = set()
        if existing and "metadatas" in existing and existing["metadatas"]:
            for metadata in existing["metadatas"]:
                if isinstance(metadata, dict) and metadata is not None:
                    doc_hash = metadata.get("doc_hash")
                    if doc_hash:
                        existing_hashes.add(doc_hash)
        print__chromadb_debug(
            f"📊 {CREATE_CHROMADB_ID}: Found {len(existing_hashes)} existing documents in ChromaDB."
        )

        # Filter out existing documents
        new_indices = [
            i for i, doc_hash in enumerate(hashes) if doc_hash not in existing_hashes
        ]
        new_texts = [texts[i] for i in new_indices]
        new_selections = [selections[i] for i in new_indices]
        new_hashes = [hashes[i] for i in new_indices]

        if not new_texts:
            print__chromadb_debug(f"⚠️ {CREATE_CHROMADB_ID}: No new documents to add.")
            return collection

        print__chromadb_debug(
            f"🔄 {CREATE_CHROMADB_ID}: Processing {len(new_texts)} new documents."
        )

        # Process documents in batches to handle token limits
        BATCH_SIZE = 1  # Process one document at a time
        total_batches = (len(new_texts) + BATCH_SIZE - 1) // BATCH_SIZE

        # Use tqdm for progress tracking
        with tqdm_module.tqdm(
            total=len(new_texts),
            desc="Processing documents",
            leave=True,
            ncols=100,
            bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
            file=sys.stdout,
        ) as pbar:
            for batch_idx in range(total_batches):
                start_idx = batch_idx * BATCH_SIZE
                end_idx = min((batch_idx + 1) * BATCH_SIZE, len(new_texts))

                batch_texts = new_texts[start_idx:end_idx]
                batch_selections = new_selections[start_idx:end_idx]
                batch_hashes = new_hashes[start_idx:end_idx]

                try:
                    # Print batch info for debugging
                    print__chromadb_debug(
                        f"📦 \n{CREATE_CHROMADB_ID}: Processing batch {batch_idx + 1}/{total_batches}"
                    )

                    # Process each document in the batch
                    for text, selection_code, doc_hash in zip(
                        batch_texts, batch_selections, batch_hashes
                    ):
                        # Count tokens and split if necessary
                        token_count = num_tokens_from_string(text)
                        print__chromadb_debug(
                            f"📊 {CREATE_CHROMADB_ID}: - {selection_code}: {len(text)} characters, {token_count} tokens"
                        )

                        # Split text if it exceeds token limit
                        text_chunks = split_text_by_tokens(text)

                        if len(text_chunks) > 1:
                            print__chromadb_debug(
                                f"✂️ {CREATE_CHROMADB_ID}: - Split {selection_code} into {len(text_chunks)} chunks"
                            )

                        # Process each chunk
                        for chunk_idx, chunk in enumerate(text_chunks):
                            try:
                                # Generate embedding for chunk
                                response = embedding_client.embeddings.create(
                                    input=[chunk], model=deployment
                                )
                                embedding = response.data[0].embedding

                                # Create unique ID and metadata for this chunk
                                chunk_id = str(uuid4())
                                chunk_metadata = {
                                    "selection": selection_code,
                                    "doc_hash": doc_hash,
                                    "chunk_index": chunk_idx,
                                    "total_chunks": len(text_chunks),
                                }

                                # Add chunk to collection
                                collection.add(
                                    documents=[chunk],
                                    embeddings=[embedding],
                                    ids=[chunk_id],
                                    metadatas=[chunk_metadata],
                                )

                                metrics.processed_docs += 1

                            except Exception as e:
                                handle_processing_error(
                                    e, f"{selection_code}_chunk_{chunk_idx}", metrics
                                )
                                continue

                    pbar.update(len(batch_texts))

                except Exception as e:
                    # Handle batch processing errors
                    for sel in batch_selections:
                        handle_processing_error(e, sel, metrics)
                    pbar.update(
                        len(batch_texts)
                    )  # Update progress even for failed batches
                    continue

        # Calculate and display final processing statistics
        metrics.update_processing_time()

        print__chromadb_debug(
            f"\nProcessing completed in {metrics.total_processing_time:.2f} seconds:"
        )
        print__chromadb_debug(f"- Total documents: {len(new_texts)}")
        print__chromadb_debug(f"- Successfully processed: {metrics.processed_docs}")
        print__chromadb_debug(f"- Failed: {metrics.failed_docs}")
        print__chromadb_debug(
            f"- Average time per document: {metrics.total_processing_time/max(1,metrics.processed_docs):.2f} seconds"
        )
        print__chromadb_debug(
            f"- Success rate: {metrics.to_dict()['success_rate']:.1f}%"
        )

        # Display failed records if any
        if metrics.failed_docs > 0:
            print__chromadb_debug("\nFailed Records:")
            print__chromadb_debug("=" * 50)
            for selection_code, error in metrics.failed_records:
                print__chromadb_debug(f"- {selection_code}: {error}")
            print__chromadb_debug("=" * 50)
            print__chromadb_debug(f"Total failed records: {metrics.failed_docs}")

            # Display failed selection codes and their description lengths
            print__chromadb_debug(
                "\nFailed Selection Codes and Description Lengths (sorted by length):"
            )
            print__chromadb_debug("=" * 50)
            # Get failed selection codes and their description lengths
            failed_selection_lengths = []
            for failed_code, _ in metrics.failed_records:
                # Find the corresponding text for this selection code
                for sel, text in zip(selections, texts):
                    if sel == failed_code:
                        failed_selection_lengths.append((failed_code, len(text)))
                        break
            # Sort by length in descending order
            failed_selection_lengths.sort(key=lambda x: x[1], reverse=True)
            # Display the results
            for selection_code, length in failed_selection_lengths:
                print__chromadb_debug(f"- {selection_code}: {length} characters")
            print__chromadb_debug("=" * 50)

        return collection

    except Exception as e:
        print__chromadb_debug(
            f"❌ {CREATE_CHROMADB_ID}: Error in upsert_documents_to_chromadb: {str(e)}"
        )
        raise


def get_chromadb_collection(
    collection_name: str,
    chroma_db_path: str,
    embedding_model_name: str = "text-embedding-3-small_mimi",
):
    """Return a direct ChromaDB collection instance for the given collection.

    This function supports both cloud and local ChromaDB based on CHROMA_USE_CLOUD env var.
    """
    from metadata.chromadb_client_factory import get_chromadb_client

    client = get_chromadb_client(
        local_path=chroma_db_path, collection_name=collection_name
    )
    collection = client.get_collection(name=collection_name)
    return collection


def similarity_search_chromadb(
    collection,
    embedding_client,
    query: str,
    embedding_model_name: str = "text-embedding-3-small_mimi",
    k: int = 3,
):
    """Perform a pure embedding-based similarity search using ChromaDB's .query method."""
    query_embedding = (
        embedding_client.embeddings.create(input=[query], model=embedding_model_name)
        .data[0]
        .embedding
    )
    print__chromadb_debug(
        f"Generated query embedding with {len(query_embedding)} dimensions for model {embedding_model_name}"
    )
    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=k,
        include=["documents", "metadatas", "distances"],
    )
    return results


def write_search_comparison_excel(query, semantic_results, hybrid_results, path):
    """
    Write a comprehensive comparison Excel file showing the agent's actual workflow:
    1. Pure semantic search (baseline)
    2. Hybrid search (semantic + BM25)
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Search_Comparison"
    ws.append(
        [
            "Query",
            "Rank_Semantic",
            "Rank_Hybrid",
            "Text_Preview",
            "Selection_Code",
            "Semantic_Score",
            "Hybrid_Score",
        ]
    )

    # Build lookup tables for fast access
    semantic_lookup = {}
    if semantic_results and "documents" in semantic_results:
        semantic_docs = semantic_results["documents"][0]
        semantic_metas = semantic_results["metadatas"][0]
        semantic_scores = semantic_results["distances"][0]
        for i, (doc, meta, distance) in enumerate(
            zip(semantic_docs, semantic_metas, semantic_scores)
        ):
            key = meta.get("selection") if meta else doc[:50]
            similarity_score = 1 - (distance / 2)  # Convert distance to similarity
            semantic_lookup[key] = (i + 1, similarity_score, doc)

    hybrid_lookup = {}
    for i, result in enumerate(hybrid_results, 1):
        key = result["metadata"].get("selection", "unknown")
        hybrid_lookup[key] = (i, result.get("score", 0), result["document"])

    # Get all unique selection codes
    all_keys = set(semantic_lookup.keys()) | set(hybrid_lookup.keys())

    # Collect all rows
    rows = []
    for key in all_keys:
        rank_sem, sem_score, sem_text = semantic_lookup.get(key, (None, None, None))
        rank_hyb, hyb_score, hyb_text = hybrid_lookup.get(key, (None, None, None))

        # Use the best available text preview
        text_preview = sem_text or hyb_text or ""
        if text_preview:
            text_preview = (
                text_preview[:200].replace("\n", " ") + "..."
                if len(text_preview) > 200
                else text_preview
            )

        rows.append(
            [
                query,
                rank_sem,
                rank_hyb,
                text_preview,
                key,
                sem_score,
                hyb_score,
            ]
        )

    # Sort by hybrid rank (since that's what feeds into reranking)
    rows_sorted = sorted(
        rows, key=lambda r: (r[2] is None, r[2] if r[2] is not None else float("inf"))
    )

    for row in rows_sorted:
        ws.append(row)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(str(path))


# ==============================================================================
# SCRIPT ENTRY POINT
# ==============================================================================
if __name__ == "__main__":
    # When running this script directly, ALWAYS use local ChromaDB
    # This ensures data loading always happens locally, regardless of .env settings
    original_cloud_setting = os.environ.get("CHROMA_USE_CLOUD")
    os.environ["CHROMA_USE_CLOUD"] = "false"

    print(f"🚀 {CREATE_CHROMADB_ID}: Starting ChromaDB creation and loading script...")
    print(
        f"🏠 {CREATE_CHROMADB_ID}: FORCING LOCAL MODE for data loading (original setting: {original_cloud_setting})"
    )
    print(f"📁 {CREATE_CHROMADB_ID}: SQLite DB path: {SQLITE_DB_PATH}")
    print(f"📁 {CREATE_CHROMADB_ID}: ChromaDB path: {CHROMA_DB_PATH}")
    print(f"📊 {CREATE_CHROMADB_ID}: SQLite DB exists: {SQLITE_DB_PATH.exists()}")

    try:
        # Initialize and populate ChromaDB
        print(f"🔄 {CREATE_CHROMADB_ID}: Calling upsert_documents_to_chromadb()...")
        collection = upsert_documents_to_chromadb()
        if collection is None:
            sys.exit(1)

        embedding_client = get_azure_embedding_model()

        # Test query - same as used in agent
        # QUERY = "Jaký je počet obytných domů vlastněných bytovými družstvy?"
        # QUERY = "Compare married and divorced people in the Czech Republic"
        # QUERY = "Kolik lidi zije na Marzu?"
        # QUERY = "kolik bylo svateb v poslednich letech?"
        QUERY = "Kolik svateb se konalo v posledních letech?"
        k = 60

        print(f"\n🔍 Testing Agent Workflow with query: '{QUERY}'")
        print(
            f"📊 Requesting top {k} results from each method (matching agent workflow)"
        )
        print("=" * 80)

        # Step 1: Pure semantic search (baseline comparison)
        print(f"\n[1/3] Pure Semantic Search (baseline)")
        semantic_results = similarity_search_chromadb(
            collection=collection,
            embedding_client=embedding_client,
            query=QUERY,
            embedding_model_name="text-embedding-3-small_mimi",
            k=k,
        )
        print(f"✅ Retrieved {len(semantic_results['documents'][0])} semantic results")
        for i, (doc, meta, distance) in enumerate(
            zip(
                semantic_results["documents"][0][:5],
                semantic_results["metadatas"][0][:5],
                semantic_results["distances"][0][:5],
            ),
            1,
        ):
            selection = (
                meta.get("selection")
                if isinstance(meta, dict) and meta is not None
                else "N/A"
            )
            similarity = 1 - (distance / 2)
            print(f"  #{i}: {selection} | Similarity: {similarity:.4f}")

        # Step 2: Hybrid search (agent's first step)
        print(f"\n[2/3] Hybrid Search (agent workflow)")
        hybrid_results = hybrid_search(collection, QUERY, n_results=k)
        print(f"✅ Retrieved {len(hybrid_results)} hybrid results")
        for i, result in enumerate(hybrid_results[:5], 1):
            selection = result["metadata"].get("selection", "unknown")
            score = result.get("score", 0)
            semantic_score = result.get("semantic_score", 0)
            bm25_score = result.get("bm25_score", 0)
            source = result.get("source", "unknown")
            print(
                f"  #{i}: {selection} | Score: {score:.6f} (sem: {semantic_score:.3f}, bm25: {bm25_score:.3f}, src: {source})"
            )

        # Display top hybrid results
        print(f"\n🎯 Final Results (top 5 from hybrid search):")
        for i, result in enumerate(hybrid_results[:5], 1):
            selection = result["metadata"].get("selection", "N/A")
            score = result.get("score", 0)
            print(f"  #{i}: {selection} | Score: {score:.6f}")

        # Final selection codes (what agent would use)
        print(f"\n🎯 Final Agent Selection Codes (top 3 above threshold 0.0005):")
        SIMILARITY_THRESHOLD = 0.0005
        final_selections = []
        for result in hybrid_results:
            selection_code = result["metadata"].get("selection")
            score = result.get("score", 0)
            if selection_code and score >= SIMILARITY_THRESHOLD:
                final_selections.append(selection_code)
        final_top3 = final_selections[:3]
        print(f"  {final_top3}")

        # Excel comparison output
        print(f"\n📊 Generating Excel comparison...")
        debug_xlsx_path = BASE_DIR / "metadata" / "search_comparison_for_debug.xlsx"
        try:
            write_search_comparison_excel(
                QUERY, semantic_results, hybrid_results, debug_xlsx_path
            )
            print(f"✅ Excel comparison written to: {debug_xlsx_path}")
        except Exception as e:
            print(f"❌ Error writing Excel comparison: {e}")

        print(f"\n🎉 Agent workflow test completed successfully!")
        print(f"📈 Workflow: Semantic → Hybrid → Top 3 selections")

    except KeyboardInterrupt:
        print__chromadb_debug(f"{CREATE_CHROMADB_ID}: Process interrupted by user")
        sys.exit(1)
    except Exception as e:
        print__chromadb_debug(f"❌ {CREATE_CHROMADB_ID}: Unexpected error: {str(e)}")
        sys.exit(1)
    finally:
        # Restore original cloud setting
        if original_cloud_setting is not None:
            os.environ["CHROMA_USE_CLOUD"] = original_cloud_setting
        elif "CHROMA_USE_CLOUD" in os.environ:
            del os.environ["CHROMA_USE_CLOUD"]
