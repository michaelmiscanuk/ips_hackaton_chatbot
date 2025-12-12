"""
ChromaDB Manager for Chatbot with Hybrid Search

This module manages ChromaDB operations for the chatbot, including:
- Loading data from CSV files
- Hybrid search (semantic + BM25)
- Document deduplication

Key Features:
- CSV data ingestion
- Hybrid search combining semantic and BM25 approaches
- Token counting and validation
- Comprehensive testing and benchmarking
"""

# ==============================================================================
# IMPORTS
# ==============================================================================
import os
import sys
import hashlib
import logging
import math
import time
import asyncio
import pandas as pd
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple, Optional
from uuid import uuid4
import tiktoken
import tqdm as tqdm_module

# Third-party imports
import chromadb
from langchain_core.documents import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Load environment variables
from dotenv import load_dotenv

load_dotenv()

# Add backend to path
try:
    BACKEND_DIR = Path(__file__).resolve().parents[1]
except NameError:
    BACKEND_DIR = Path(os.getcwd()).parents[0]

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

# Import BM25 if available
try:
    from rank_bm25 import BM25Okapi

    print("[OK] rank_bm25 is available. BM25 search will be enabled.")
except ImportError:
    print("[WARNING] rank_bm25 not available. BM25 search will be disabled.")
    BM25Okapi = None

# Import local modules
from openai import AzureOpenAI
from langchain_openai import AzureOpenAIEmbeddings
from langchain_chroma import Chroma
from src.config.models import get_embedding_model, get_embedding_model_name

# Import embedding model configuration
try:
    from langchain_ollama import OllamaEmbeddings
except ImportError:
    OllamaEmbeddings = None

# ==============================================================================
# CONSTANTS & CONFIGURATION
# ==============================================================================
# Use BACKEND_DIR (backend/) to locate data
BASE_DIR = BACKEND_DIR.parent  # Root dir
CSV_PATH = BACKEND_DIR / "data" / "sample0.csv"
MAX_TOKENS = 8190


def get_chromadb_dir():
    """Get ChromaDB directory with embedding model suffix."""
    model_name = get_embedding_model_name()
    return BACKEND_DIR / "data" / f"chroma_db_{model_name}"


def get_collection_name():
    """Get collection name with embedding model suffix."""
    model_name = get_embedding_model_name()
    # Sanitize model name for collection names
    model_name = (
        model_name.replace("/", "-")
        .replace("\\", "-")
        .replace(":", "-")
        .replace(".", "_")
    )
    return f"chatbot_{model_name}"


CHROMA_DB_DIR = get_chromadb_dir()

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================
def get_azure_embedding_model():
    """Get an instance of Azure OpenAI Embedding model."""
    return AzureOpenAI(
        api_version="2024-12-01-preview",
        azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
        api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    )


def get_langchain_azure_embedding_model(deployment_name="text-embedding-3-small_mimi"):
    """Get a LangChain AzureOpenAIEmbeddings instance."""
    return AzureOpenAIEmbeddings(
        model="text-embedding-3-small",
        azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
        api_key=os.getenv("AZURE_OPENAI_API_KEY"),
        openai_api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview"),
        deployment=deployment_name,
    )


# get_embedding_model is now imported from src.config.models
# We remove the local definition to avoid duplication and ensure consistency


def get_document_hash(text: str) -> str:
    """Generate MD5 hash for a document text."""
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def num_tokens_from_string(string: str, encoding_name: str = "cl100k_base") -> int:
    """Count the number of tokens in a string."""
    encoding = tiktoken.get_encoding(encoding_name)
    return len(encoding.encode(string))


def split_text_by_tokens(text: str, max_tokens: int = MAX_TOKENS) -> List[str]:
    """Split text into chunks that don't exceed the token limit."""
    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(text)
    total_tokens = len(tokens)

    if total_tokens <= max_tokens:
        return [text]

    num_chunks = math.ceil(float(total_tokens) / max_tokens)
    chunks = []

    for i in range(num_chunks):
        start_idx = i * max_tokens
        end_idx = min((i + 1) * max_tokens, total_tokens)
        chunk_tokens = tokens[start_idx:end_idx]
        chunk_text = encoding.decode(chunk_tokens)
        chunks.append(chunk_text)

    return chunks


def normalize_czech_text(text: str) -> str:
    """Advanced Czech text normalization for better search matching."""
    if not text:
        return text

    text = text.lower()

    czech_diacritics_map = {
        "á": "a",
        "č": "c",
        "ď": "d",
        "é": "e",
        "ě": "e",
        "í": "i",
        "ň": "n",
        "ó": "o",
        "ř": "r",
        "š": "s",
        "ť": "t",
        "ú": "u",
        "ů": "u",
        "ý": "y",
        "ž": "z",
    }

    ascii_text = text
    for diacritic, ascii_char in czech_diacritics_map.items():
        ascii_text = ascii_text.replace(diacritic, ascii_char)

    if ascii_text != text:
        return f"{text} {ascii_text}"
    return text


# ==============================================================================
# DATA LOADING
# ==============================================================================
@dataclass
class Metrics:
    """Track processing metrics."""

    start_time: datetime = field(default_factory=datetime.now)
    processed_docs: int = 0
    failed_docs: int = 0
    failed_records: List[Tuple[str, str]] = field(default_factory=list)
    total_processing_time: float = 0.0

    def update_processing_time(self):
        """Update total processing time."""
        self.total_processing_time = (datetime.now() - self.start_time).total_seconds()

    def to_dict(self) -> Dict[str, Any]:
        """Convert metrics to dictionary."""
        success_rate = (
            (self.processed_docs / (self.processed_docs + self.failed_docs) * 100)
            if (self.processed_docs + self.failed_docs) > 0
            else 0
        )
        return {
            "processed_docs": self.processed_docs,
            "failed_docs": self.failed_docs,
            "success_rate": success_rate,
            "total_time": self.total_processing_time,
        }


async def process_csv_row(row):
    """Process a single CSV row: format and create document."""
    # Handle NaN/float values by converting to string
    subject = str(row.get("subject", "")) if pd.notna(row.get("subject")) else ""
    body = str(row.get("body", "")) if pd.notna(row.get("body")) else ""
    answer = str(row.get("answer", "")) if pd.notna(row.get("answer")) else ""
    row_type = str(row.get("type", "")) if pd.notna(row.get("type")) else ""
    language = str(row.get("language", "")) if pd.notna(row.get("language")) else ""

    text_content = f"Subject: {subject}\nBody: {body}\nAnswer: {answer}"

    subject_preview = subject[:30] if len(subject) > 30 else subject
    logger.info(f"Processing row: {subject_preview}...")

    metadata = {
        "subject": subject,
        "type": row_type,
        "language": language,
        "original_text": text_content,
        "hash": get_document_hash(text_content),
    }

    return Document(page_content=text_content, metadata=metadata)


async def load_documents_from_csv() -> List[Document]:
    """Load and process documents from CSV file."""
    logger.info(f"Reading data from {CSV_PATH}...")

    if not CSV_PATH.exists():
        raise FileNotFoundError(f"CSV file not found: {CSV_PATH}")

    df = pd.read_csv(CSV_PATH)
    logger.info(f"Found {len(df)} records in CSV")

    documents = []
    for index, row in df.iterrows():
        try:
            doc = await process_csv_row(row)
            documents.append(doc)
        except Exception as e:
            logger.error(f"Error processing row {index}: {e}")
            import traceback

            logger.error(traceback.format_exc())
            raise

    logger.info(f"Prepared {len(documents)} documents")
    return documents


# ==============================================================================
# HYBRID SEARCH IMPLEMENTATION
# ==============================================================================
def hybrid_search(
    collection, query_text: str, n_results: int = 60, rare_terms: Set[str] = None
) -> List[Dict]:
    """
    Hybrid search combining semantic and BM25 approaches.

    Semantic search gets higher weight (0.85) as the primary method.
    BM25 search gets lower weight (0.15) for exact matches.
    """
    logger.info(f"Hybrid search for query: '{query_text}'")

    try:
        # Step 1: Semantic search
        # Get embedding model (Ollama or Azure based on config)
        embedding_model = get_embedding_model()

        # Generate query embedding
        query_embedding = embedding_model.embed_query(query_text)

        semantic_results = collection.query(
            query_embeddings=[query_embedding],
            n_results=n_results,
            include=["documents", "metadatas", "distances"],
        )

        # Convert to normalized scores
        semantic_docs = {}
        if semantic_results and "documents" in semantic_results:
            docs_list = semantic_results["documents"][0]
            metas_list = semantic_results["metadatas"][0]
            distances = semantic_results["distances"][0]

            # Normalize distances to [0, 1] similarity scores
            max_dist = max(distances) if distances else 1
            min_dist = min(distances) if distances else 0
            dist_range = max_dist - min_dist if max_dist != min_dist else 1

            for doc, meta, dist in zip(docs_list, metas_list, distances):
                norm_score = 1 - ((dist - min_dist) / dist_range)
                doc_id = meta.get("subject", doc[:50])
                semantic_docs[doc_id] = {
                    "document": doc,
                    "metadata": meta,
                    "semantic_score": norm_score,
                }

        # Step 2: BM25 search (if available)
        bm25_docs = {}
        if BM25Okapi:
            try:
                # Get all documents for BM25
                all_docs_result = collection.get(include=["documents", "metadatas"])
                if all_docs_result and "documents" in all_docs_result:
                    corpus = all_docs_result["documents"]
                    corpus_metas = all_docs_result["metadatas"]

                    # Tokenize corpus
                    tokenized_corpus = [doc.lower().split() for doc in corpus]
                    bm25 = BM25Okapi(tokenized_corpus)

                    # Get BM25 scores
                    tokenized_query = query_text.lower().split()
                    bm25_scores = bm25.get_scores(tokenized_query)

                    # Normalize BM25 scores
                    max_score = max(bm25_scores) if len(bm25_scores) > 0 else 1
                    min_score = min(bm25_scores) if len(bm25_scores) > 0 else 0
                    score_range = max_score - min_score if max_score != min_score else 1

                    # Get top BM25 results
                    top_indices = sorted(
                        range(len(bm25_scores)),
                        key=lambda i: bm25_scores[i],
                        reverse=True,
                    )[:n_results]

                    for idx in top_indices:
                        if bm25_scores[idx] > 0:
                            norm_score = (bm25_scores[idx] - min_score) / score_range
                            doc = corpus[idx]
                            meta = corpus_metas[idx]
                            doc_id = meta.get("subject", doc[:50])
                            bm25_docs[doc_id] = {
                                "document": doc,
                                "metadata": meta,
                                "bm25_score": norm_score,
                            }

            except Exception as e:
                logger.error(f"BM25 search error: {e}")

        # Step 3: Combine results with weighted scoring
        SEMANTIC_WEIGHT = 0.85
        BM25_WEIGHT = 0.15

        combined_results = {}

        # Add semantic results
        for doc_id, doc_data in semantic_docs.items():
            combined_results[doc_id] = {
                "document": doc_data["document"],
                "metadata": doc_data["metadata"],
                "semantic_score": doc_data["semantic_score"],
                "bm25_score": 0.0,
                "score": doc_data["semantic_score"] * SEMANTIC_WEIGHT,
                "source": "semantic",
            }

        # Add/update with BM25 results
        for doc_id, doc_data in bm25_docs.items():
            if doc_id in combined_results:
                combined_results[doc_id]["bm25_score"] = doc_data["bm25_score"]
                combined_results[doc_id]["score"] += (
                    doc_data["bm25_score"] * BM25_WEIGHT
                )
                combined_results[doc_id]["source"] = "hybrid"
            else:
                combined_results[doc_id] = {
                    "document": doc_data["document"],
                    "metadata": doc_data["metadata"],
                    "semantic_score": 0.0,
                    "bm25_score": doc_data["bm25_score"],
                    "score": doc_data["bm25_score"] * BM25_WEIGHT,
                    "source": "bm25",
                }

        # Sort by combined score
        results = sorted(
            combined_results.values(), key=lambda x: x["score"], reverse=True
        )

        logger.info(f"Hybrid search returned {len(results)} results")
        return results[:n_results]

    except Exception as e:
        logger.error(f"Hybrid search error: {e}")
        return []


# ==============================================================================
# CHROMADB OPERATIONS
# ==============================================================================
async def upsert_documents_to_chromadb(
    deployment: str = None,  # Kept for backward compatibility but ignored
    collection_name: str = None,  # Auto-generated if not provided
) -> chromadb.Collection:
    """Load documents from CSV and add to ChromaDB.

    Args:
        deployment: Deprecated, kept for backward compatibility
        collection_name: Collection name. If None, auto-generated with embedding model suffix
    """
    if collection_name is None:
        collection_name = get_collection_name()

    # Update CHROMA_DB_DIR to current model's directory
    chroma_db_dir = get_chromadb_dir()
    """Load documents from CSV and add to ChromaDB."""
    metrics = Metrics()

    logger.info("=" * 70)
    logger.info("Starting ChromaDB Upsert Process")
    logger.info("=" * 70)

    try:
        # Load documents from CSV
        documents = await load_documents_from_csv()

        if not documents:
            raise ValueError("No documents loaded from CSV")

        # Initialize embedding model (Ollama or Azure based on config)
        embedding_model = get_embedding_model()

        # Initialize ChromaDB
        logger.info(f"Creating ChromaDB at {chroma_db_dir}...")
        chroma_db_dir.mkdir(parents=True, exist_ok=True)

        client = chromadb.PersistentClient(path=str(chroma_db_dir))

        # Delete existing collection if it exists
        try:
            client.delete_collection(name=collection_name)
            logger.info(f"Deleted existing collection: {collection_name}")
        except:
            pass

        # Create new collection
        collection = client.create_collection(
            name=collection_name,
            metadata={"hnsw:space": "cosine"},
        )

        # Process documents
        logger.info(f"Processing {len(documents)} documents...")

        # Increase batch size for better GPU utilization
        # GPU memory is underutilized, so we can process larger batches
        batch_size = 20  # Increased to 20 for better GPU utilization

        # Prepare all batches
        batches = []
        for i in range(0, len(documents), batch_size):
            batch = documents[i : i + batch_size]
            texts = [doc.page_content for doc in batch]
            metadatas = [doc.metadata for doc in batch]
            ids = [str(uuid4()) for _ in batch]
            batches.append((texts, metadatas, ids))

        logger.info(f"Processing {len(batches)} batches with batch_size={batch_size}")

        # Process batches with retry logic for transient Ollama connection errors
        for batch_idx, (texts, metadatas, ids) in enumerate(
            tqdm_module.tqdm(batches, desc="Adding documents")
        ):
            # Retry logic for connection errors
            max_retries = 3
            retry_delay = 2  # seconds

            for retry in range(max_retries):
                try:
                    # Generate embeddings for the entire batch at once
                    # Ollama will handle this efficiently with GPU if available
                    embeddings = embedding_model.embed_documents(texts)

                    # Validate embeddings format
                    if not isinstance(embeddings, list):
                        logger.error(f"Embeddings is not a list: {type(embeddings)}")
                        raise TypeError(
                            f"Expected list of embeddings, got {type(embeddings)}"
                        )

                    # Ensure embeddings is a list of lists (vectors)
                    validated_embeddings = []
                    for idx, emb in enumerate(embeddings):
                        if isinstance(emb, (list, tuple)):
                            validated_embeddings.append(list(emb))
                        elif hasattr(emb, "__iter__") and not isinstance(emb, str):
                            validated_embeddings.append(list(emb))
                        else:
                            logger.error(
                                f"Invalid embedding at index {idx}: {type(emb)}"
                            )
                            raise TypeError(
                                f"Invalid embedding type at index {idx}: {type(emb)}"
                            )

                    # Add to collection
                    collection.add(
                        documents=texts,
                        embeddings=validated_embeddings,
                        metadatas=metadatas,
                        ids=ids,
                    )

                    metrics.processed_docs += len(texts)
                    break  # Success, exit retry loop

                except Exception as e:
                    error_msg = str(e)
                    # Check if it's a transient connection error
                    if (
                        "connection" in error_msg.lower()
                        or "500" in error_msg
                        or retry < max_retries - 1
                    ):
                        if retry < max_retries - 1:
                            wait_time = retry_delay * (2**retry)  # Exponential backoff
                            logger.warning(
                                f"Batch {batch_idx} failed (attempt {retry + 1}/{max_retries}), retrying in {wait_time}s: {e}"
                            )
                            time.sleep(wait_time)
                        else:
                            logger.error(
                                f"Batch {batch_idx} failed after {max_retries} attempts: {e}"
                            )
                            import traceback

                            logger.error(traceback.format_exc())
                            metrics.failed_docs += len(texts)
                    else:
                        # Non-retryable error
                        logger.error(f"Error processing batch {batch_idx}: {e}")
                        import traceback

                        logger.error(traceback.format_exc())
                        metrics.failed_docs += len(texts)
                        break

        metrics.update_processing_time()

        logger.info("=" * 70)
        logger.info("ChromaDB Upsert Completed")
        logger.info(f"Processed: {metrics.processed_docs} documents")
        logger.info(f"Failed: {metrics.failed_docs} documents")
        logger.info(f"Time: {metrics.total_processing_time:.2f} seconds")
        logger.info("=" * 70)

        return collection

    except Exception as e:
        logger.error(f"Error in upsert_documents_to_chromadb: {e}")
        raise


def get_chromadb_collection(
    collection_name: str = None,
) -> chromadb.Collection:
    """Get an existing ChromaDB collection.

    Args:
        collection_name: Collection name. If None, auto-generated with embedding model suffix

    Note:
        Falls back to 'chatbot_collection' if the model-specific collection doesn't exist,
        for backward compatibility with existing ChromaDBs.
    """
    if collection_name is None:
        collection_name = get_collection_name()

    chroma_db_dir = get_chromadb_dir()
    client = chromadb.PersistentClient(path=str(chroma_db_dir))

    # Try to get the collection with the new name first
    try:
        return client.get_collection(name=collection_name)
    except Exception:
        # Fallback to old collection name for backward compatibility
        logger.info(
            f"Collection '{collection_name}' not found, trying fallback 'chatbot_collection'"
        )
        return client.get_collection(name="chatbot_collection")


def similarity_search_chromadb(
    collection,
    embedding_client,
    query: str,
    deployment_name: str = None,  # Deprecated - kept for backward compatibility
    k: int = 3,
):
    """Perform pure embedding-based similarity search."""
    # Use the configured embedding model instead of Azure-specific client
    embedding_model = get_embedding_model()
    query_embedding = embedding_model.embed_query(query)

    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=k,
        include=["documents", "metadatas", "distances"],
    )
    return results


# ==============================================================================
# TESTING AND BENCHMARKING
# ==============================================================================
def write_search_comparison_excel(query, semantic_results, hybrid_results, path):
    """Write search comparison to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Search_Comparison"

    ws.append(
        [
            "Query",
            "Rank_Semantic",
            "Rank_Hybrid",
            "Text_Preview",
            "Subject",
            "Semantic_Score",
            "Hybrid_Score",
        ]
    )

    # Build lookup tables
    semantic_lookup = {}
    if semantic_results and "documents" in semantic_results:
        # Safely get documents, metadatas, and distances
        docs = (
            semantic_results.get("documents", [[]])[0]
            if isinstance(semantic_results.get("documents", [[]]), list)
            else []
        )
        metas = (
            semantic_results.get("metadatas", [[]])[0]
            if isinstance(semantic_results.get("metadatas", [[]]), list)
            else []
        )
        distances = (
            semantic_results.get("distances", [[]])[0]
            if isinstance(semantic_results.get("distances", [[]]), list)
            else []
        )

        for i, (doc, meta, distance) in enumerate(zip(docs, metas, distances)):
            key = meta.get("subject", doc[:50]) if isinstance(meta, dict) else doc[:50]
            similarity = 1 - (distance / 2) if isinstance(distance, (int, float)) else 0
            semantic_lookup[key] = (i + 1, similarity, doc)

    hybrid_lookup = {}
    for i, result in enumerate(hybrid_results, 1):
        key = result["metadata"].get("subject", "unknown")
        hybrid_lookup[key] = (i, result.get("score", 0), result["document"])

    # Combine all keys
    all_keys = set(semantic_lookup.keys()) | set(hybrid_lookup.keys())

    # Collect rows
    rows = []
    for key in all_keys:
        rank_sem, sem_score, sem_text = semantic_lookup.get(key, (None, None, None))
        rank_hyb, hyb_score, hyb_text = hybrid_lookup.get(key, (None, None, None))

        text_preview = sem_text or hyb_text or ""
        if text_preview:
            text_preview = (
                text_preview[:200].replace("\n", " ") + "..."
                if len(text_preview) > 200
                else text_preview
            )

        rows.append(
            [
                query,
                rank_sem,
                rank_hyb,
                text_preview,
                key,
                sem_score,
                hyb_score,
            ]
        )

    # Sort by hybrid rank
    rows_sorted = sorted(
        rows, key=lambda r: (r[2] is None, r[2] if r[2] is not None else float("inf"))
    )

    for row in rows_sorted:
        ws.append(row)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(str(path))
    logger.info(f"Excel comparison saved to {path}")


# ==============================================================================
# SCRIPT ENTRY POINT
# ==============================================================================
if __name__ == "__main__":
    import sys

    logger.info("Starting ChromaDB Manager Script")

    async def main():
        # Create and populate ChromaDB
        collection = await upsert_documents_to_chromadb()

        # Test queries
        QUERY = "How do I reset my password?"
        k = 10

        logger.info(f"\n{'=' * 70}")
        logger.info(f"Testing with query: '{QUERY}'")
        logger.info(f"{'=' * 70}\n")

        # Step 1: Pure semantic search
        logger.info("[1/3] Pure Semantic Search")
        # We pass None for embedding_client as it's ignored in favor of get_embedding_model()
        semantic_results = similarity_search_chromadb(
            collection=collection,
            embedding_client=None,
            query=QUERY,
            k=k,
        )
        logger.info(f"[OK] Retrieved {len(semantic_results['documents'][0])} results")

        # Step 2: Hybrid search
        logger.info("\n[2/3] Hybrid Search")
        hybrid_results = hybrid_search(collection, QUERY, n_results=k)
        logger.info(f"[OK] Retrieved {len(hybrid_results)} results")

        # Display top results from hybrid search
        logger.info("\n[RESULTS] Top 3 Results:")
        for i, result in enumerate(hybrid_results[:3], 1):
            subject = result["metadata"].get("subject", "N/A")
            score = result.get("score", 0)
            logger.info(f"  #{i}: {subject} | Score: {score:.4f}")

        # Save comparison
        logger.info("\n[EXPORT] Saving comparison to Excel...")
        excel_path = BASE_DIR / "backend" / "data" / "search_comparison.xlsx"
        write_search_comparison_excel(
            QUERY, semantic_results, hybrid_results, excel_path
        )

        logger.info("\n[SUCCESS] Test completed successfully!")

    asyncio.run(main())
